import openpyxl
from openpyxl.utils.cell import column_index_from_string
import datetime
# 現在の日付（年月日）を取得、変数todayに代入する
today = datetime.date.today()
# 新規フィルの作成
workbook = openpyxl.Workbook()
# シートの選択
sheet = workbook.active
# シートの列の幅を変更
sheet.column_dimensions['E'].width = 15

# セルに値を設定
sheet["B2"].value = "請求書"
sheet["B4"].value = "株式会社ABC"
sheet["B5"].value = "〒101-0022 東京都千代田区神田練塀町300"
sheet["B6"].value = "TEL:03-1234-5678 FAX:03-1234-5678"
sheet["B7"].value = "担当者名:鈴木一郎 様"
sheet["F4"].value = "No."
sheet["F5"].value = "日付"
sheet["G4"].value = "0001"
sheet["G5"].value = today.strftime('%Y/%m/%d')

header = ['商品名', '数量', '単価', '金額']

# 任意のセル範囲を開始点とするため、開始点を特定（例：C3）
start_col = 'B'
start_row = 10

# ヘッダーの記述
for i, h in enumerate(header, start=0):
    sheet.cell(row=start_row, column=column_index_from_string(start_col)+i).value = h

data = [
    ['商品A', 2, 10000, 20000],
    ['商品B', 1, 15000, 15000]
]

# データの記述
for r, row in enumerate(data, start = start_row + 1):
    for c, item in enumerate(row, start = 0):
        sheet.cell(row=r, column=column_index_from_string(start_col)+c).value = item

subtotal = data[0][3] + data[1][3]
tax = 1.1
sheet["E13"].value = subtotal
sheet["B15"].value = "小計"
sheet["E15"].value = subtotal
sheet["B16"].value = "消費税"
sheet["E16"].value = (subtotal * tax) - subtotal
sheet["B17"].value = "合計"
sheet["E17"].value = subtotal * tax

# 日付のtodayオブジェクトを、特定のフォーマットで文字列に変換する
invoice_filename = f"請求書_{today.strftime('%Y%m%d')}.xlsx"
workbook.save(invoice_filename)